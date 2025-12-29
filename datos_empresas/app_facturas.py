import streamlit as st
import pandas as pd
import openpyxl
import io
import re
import unicodedata

# --- Funciones de Limpieza y Utilidad ---

def normalize_text(text):
    """
    Elimina acentos y normaliza a mayúsculas para comparaciones robustas.
    Ej: "Condición" -> "CONDICION"
    """
    if not text: return ""
    text = str(text).upper().strip()
    # Eliminar acentos
    text = ''.join(c for c in unicodedata.normalize('NFD', text) if unicodedata.category(c) != 'Mn')
    return text

def clean_text(text):
    """Limpieza básica manteniendo acentos para visualización."""
    if text:
        return str(text).strip().upper()
    return ""

def get_float(s):
    """Convierte a float limpiando símbolos no numéricos."""
    if s is None: return 0.0
    try:
        s_clean = re.sub(r'[^\d.,-]', '', str(s))
        if not s_clean: return 0.0
        s_clean = s_clean.replace(',', '')
        return float(s_clean)
    except ValueError:
        return 0.0

def parse_month(text):
    """Convierte meses texto a número."""
    text = normalize_text(text)
    months = {
        'ENERO': '01', 'JANUARY': '01', 'JAN': '01',
        'FEBRERO': '02', 'FEBRUARY': '02', 'FEB': '02',
        'MARZO': '03', 'MARCH': '03', 'MAR': '03',
        'ABRIL': '04', 'APRIL': '04', 'APR': '04',
        'MAYO': '05', 'MAY': '05',
        'JUNIO': '06', 'JUNE': '06', 'JUN': '06',
        'JULIO': '07', 'JULY': '07', 'JUL': '07',
        'AGOSTO': '08', 'AUGUST': '08', 'AUG': '08',
        'SEPTIEMBRE': '09', 'SEPTEMBER': '09', 'SEP': '09',
        'OCTUBRE': '10', 'OCTOBER': '10', 'OCT': '10',
        'NOVIEMBRE': '11', 'NOVEMBER': '11', 'NOV': '11',
        'DICIEMBRE': '12', 'DECEMBER': '12', 'DEC': '12'
    }
    for k, v in months.items():
        if k in text: return v
    if text.isdigit() and len(text) == 1: return f"0{text}"
    return text if text.isdigit() else None

# --- Lógica de Búsqueda Estructural ---

def find_label_cell(sheet, keywords):
    """
    Busca la coordenada de una etiqueta en la hoja usando texto normalizado.
    """
    keywords_norm = [normalize_text(k) for k in keywords]
    
    for row in sheet.iter_rows(min_row=1, max_row=150, max_col=50):
        for cell in row:
            val = normalize_text(cell.value)
            if not val: continue
            
            for k in keywords_norm:
                if k in val:
                    return cell.row, cell.column
    return None

def extract_value_near_label(sheet, label_row, label_col, look_down=True, look_right=True, max_steps=10):
    """Busca el valor asociado saltando vacíos."""
    directions = []
    if look_down: directions.append('below')
    if look_right: directions.append('right')
    
    for direction in directions:
        for step in range(1, max_steps + 1):
            r, c = label_row, label_col
            if direction == 'below': r += step
            elif direction == 'right': c += step
            
            try:
                cell = sheet.cell(row=r, column=c)
                val = cell.value
                val_norm = normalize_text(val)
                if val_norm:
                    # Filtro: Si parece otra etiqueta, ignorar
                    if "CLIENTE" in val_norm or "FECHA" in val_norm: continue
                    return val # Devolvemos valor original
            except:
                pass
    return None

def scan_headers_footers(sheet):
    """Busca texto en las propiedades de impresión (Header/Footer)."""
    texts = []
    try:
        props = [
            sheet.oddHeader.left, sheet.oddHeader.center, sheet.oddHeader.right,
            sheet.oddFooter.left, sheet.oddFooter.center, sheet.oddFooter.right,
            sheet.evenHeader.left, sheet.evenHeader.center, sheet.evenHeader.right,
            sheet.evenFooter.left, sheet.evenFooter.center, sheet.evenFooter.right,
            sheet.firstHeader.left, sheet.firstHeader.center, sheet.firstHeader.right,
            sheet.firstFooter.left, sheet.firstFooter.center, sheet.firstFooter.right,
        ]
        for p in props:
            if p.text: texts.append(p.text)
    except:
        pass
    return " | ".join(texts) if texts else None

def find_longest_text_in_footer(sheet, start_row, min_length=20):
    """Fuerza Bruta para encontrar textos largos al final."""
    longest_text = None
    max_len = 0
    
    for r in range(start_row, min(start_row + 50, sheet.max_row + 1)):
        for c in range(1, 25): 
            try:
                cell = sheet.cell(row=r, column=c)
                val = str(cell.value).strip() if cell.value else ""
                val_norm = normalize_text(val)
                
                # Ignorar basura común
                if any(x in val_norm for x in ["TOTAL", "PAGE", "FIRMA", "SIGNATURE", "GRACIAS", "PAGINA"]):
                    continue
                
                if len(val) > min_length and len(val) > max_len:
                    max_len = len(val)
                    longest_text = val
            except:
                continue
    return longest_text

# --- Debug Helper ---
def get_all_sheet_text(sheet):
    """Extrae todo el texto visible en celdas para diagnóstico."""
    data = []
    # Revisar celdas normales
    for r in range(1, min(sheet.max_row + 1, 200)):
        for c in range(1, min(sheet.max_column + 1, 30)):
            cell = sheet.cell(row=r, column=c)
            if cell.value:
                data.append({
                    "Ubicación": f"Fila {r}, Col {c}",
                    "Tipo": "Celda",
                    "Valor": str(cell.value)
                })
    
    # Revisar Headers/Footers
    hf_text = scan_headers_footers(sheet)
    if hf_text:
        data.append({"Ubicación": "Header/Footer", "Tipo": "Impresión", "Valor": hf_text})
        
    return pd.DataFrame(data)

# --- Procesamiento de Archivo ---

def process_file(uploaded_file):
    try:
        wb = openpyxl.load_workbook(uploaded_file, data_only=True)
        sheet = wb.active
    except Exception as e:
        return [], f"Error leyendo Excel: {str(e)}", None, False

    header_data = {}
    
    # Detección de posibles cuadros de texto (Drawings)
    has_drawings = hasattr(sheet, 'drawings') and len(sheet.drawings) > 0
    
    # Generar datos de diagnóstico
    debug_df = get_all_sheet_text(sheet)

    # 1. CLIENTE
    lbl_coords = find_label_cell(sheet, ['CLIENTE', 'CUSTOMER', 'CONSIGNEE', 'SOLD TO'])
    header_data['Cliente'] = extract_value_near_label(sheet, lbl_coords[0], lbl_coords[1]) if lbl_coords else None

    # 2. EXP
    lbl_coords = find_label_cell(sheet, ['EXP', 'INVOICE NO', 'FACTURA N'])
    header_data['Num_Exp'] = extract_value_near_label(sheet, lbl_coords[0], lbl_coords[1]) if lbl_coords else None

    # 3. FECHA
    cy = find_label_cell(sheet, ['AÑO', 'YEAR'])
    cm = find_label_cell(sheet, ['MES', 'MONTH'])
    cd = find_label_cell(sheet, ['DIA', 'DAY'])
    
    val_y = extract_value_near_label(sheet, cy[0], cy[1]) if cy else None
    val_m = extract_value_near_label(sheet, cm[0], cm[1]) if cm else None
    val_d = extract_value_near_label(sheet, cd[0], cd[1]) if cd else None
    
    if val_y and val_m and val_d:
        m_num = parse_month(str(val_m))
        header_data['Fecha'] = f"{val_d}/{m_num}/{val_y}"
    else:
        c_date = find_label_cell(sheet, ['FECHA', 'DATE'])
        if c_date:
            raw_date = extract_value_near_label(sheet, c_date[0], c_date[1])
            if hasattr(raw_date, 'strftime'):
                header_data['Fecha'] = raw_date.strftime('%d/%m/%Y')
            else:
                header_data['Fecha'] = raw_date
        else:
            header_data['Fecha'] = None

    # 4. CONDICIÓN DE VENTA - Lógica V15 (Prioridad Inversa)
    condicion_found = None
    
    # A) PRIORIDAD MÁXIMA: Buscar la frase "BAJO CONDICION" en toda la hoja (Rescue scan)
    # Esto evita que una etiqueta "Flete: Collect" oculte la verdadera condición
    for idx, row in debug_df.iterrows():
        txt = normalize_text(row['Valor'])
        # Buscamos frases que indiquen inequívocamente la condición especial
        if "BAJO CONDICION" in txt or "UNDER CONDITION" in txt or "A CONSIGNACION" in txt:
            condicion_found = row['Valor'] # Usar valor original
            break
            
    if condicion_found:
        header_data['Condicion_Venta'] = condicion_found
    else:
        # B) Si no hay frase mágica, buscar por etiqueta estándar fuerte
        lbl_coords = find_label_cell(sheet, ['CONDICION DE VENTA', 'TERMS OF SALE', 'DELIVERY TERMS', 'INCOTERM'])
        if lbl_coords:
             header_data['Condicion_Venta'] = extract_value_near_label(sheet, lbl_coords[0], lbl_coords[1], look_down=True, look_right=True)
        else:
            # C) Último recurso: Buscar TIPO FLETE (etiqueta débil)
            lbl_flete = find_label_cell(sheet, ['TIPO FLETE', 'FREIGHT TYPE'])
            if lbl_flete:
                 header_data['Condicion_Venta'] = extract_value_near_label(sheet, lbl_flete[0], lbl_flete[1], look_down=True, look_right=True)
            else:
                 header_data['Condicion_Venta'] = None


    # 5. FORMA DE PAGO
    lbl_coords = find_label_cell(sheet, ['FORMA DE PAGO', 'PAYMENT TERMS'])
    header_data['Forma_Pago'] = extract_value_near_label(sheet, lbl_coords[0], lbl_coords[1]) if lbl_coords else None

    # 6. PUERTOS
    c_emb = find_label_cell(sheet, ['PUERTO DE EMBARQUE', 'LOADING PORT', 'POL'])
    header_data['Puerto_Emb'] = extract_value_near_label(sheet, c_emb[0], c_emb[1]) if c_emb else None
    
    c_dest = find_label_cell(sheet, ['PUERTO DESTINO', 'DISCHARGING PORT', 'DESTINATION', 'POD'])
    header_data['Puerto_Dest'] = extract_value_near_label(sheet, c_dest[0], c_dest[1]) if c_dest else None

    # 7. MONEDA e INCOTERM (Global scan)
    detected_codes = {'Moneda': None, 'Incoterm': None}
    incoterm_list = ['FOB', 'CIF', 'CFR', 'EXW', 'FCA']
    currency_map = {'DÓLAR': 'USD', 'DOLAR': 'USD', 'USD': 'USD', 'EURO': 'EUR'}
    
    # 8. TIPO DE CAMBIO
    lbl_tc = find_label_cell(sheet, ['TIPO DE CAMBIO', 'TIPO CAMBIO', 'T.C.', 'EXCHANGE RATE'])
    header_data['Tipo_Cambio'] = extract_value_near_label(sheet, lbl_tc[0], lbl_tc[1]) if lbl_tc else None

    # Escaneo Global de Celdas
    for idx, row in debug_df.iterrows():
        v_orig = str(row['Valor'])
        v_norm = normalize_text(v_orig)
        
        # Detección de Incoterm (FOB, CIF)
        if not detected_codes['Incoterm']:
            for inc in incoterm_list:
                if re.search(rf"\b{inc}\b", v_norm): detected_codes['Incoterm'] = inc
        
        # Detección de Moneda (USD, DÓLAR)
        if not detected_codes['Moneda']:
            for cur, code in currency_map.items():
                if cur in v_norm: 
                    detected_codes['Moneda'] = code 
                    
    header_data['Moneda'] = detected_codes['Moneda']
    header_data['Incoterm'] = detected_codes['Incoterm']

    # --- EXTRACCIÓN DE PRODUCTOS ---
    products = []
    col_map = {}
    header_row = None
    last_row_table = 20
    
    kw_desc = ['DESCRIPCION', 'DESCRIPTION', 'MERCADERIA']
    kw_qty = ['CANTIDAD', 'QUANTITY', 'QTTY', 'PCS', 'BOXES']
    kw_price = ['PRECIO', 'PRICE', 'UNIT']
    kw_total = ['TOTAL']

    # Buscar cabecera
    for r in range(1, 100):
        row_cells = []
        for c in range(1, 30):
            val = sheet.cell(row=r, column=c).value
            row_cells.append((c, normalize_text(val)))
        
        row_txt = [x[1] for x in row_cells]
        if any(any(k in t for k in kw_desc) for t in row_txt) and any(any(k in t for k in kw_qty) for t in row_txt):
            header_row = r
            for c, v in row_cells:
                if not v: continue
                if any(k in v for k in kw_desc): col_map['Descripcion'] = c
                elif any(k in v for k in kw_qty): col_map['Cantidad'] = c
                elif any(k in v for k in kw_price): col_map['Precio Unitario'] = c
                elif any(k in v for k in kw_total) and 'TOTAL CASES' not in v and 'FOB' not in v: 
                    col_map['Total Linea'] = c
            break

    if header_row and 'Descripcion' in col_map:
        curr = header_row + 1
        empty_streak = 0
        while curr <= sheet.max_row:
            desc_val = sheet.cell(row=curr, column=col_map['Descripcion']).value
            desc_norm = normalize_text(desc_val)
            
            # Criterio de parada
            if desc_norm.startswith("TOTAL"):
                last_row_table = curr
                break
            
            qty = sheet.cell(row=curr, column=col_map['Cantidad']).value if 'Cantidad' in col_map else 0
            price = sheet.cell(row=curr, column=col_map['Precio Unitario']).value if 'Precio Unitario' in col_map else 0
            total = sheet.cell(row=curr, column=col_map['Total Linea']).value if 'Total Linea' in col_map else 0
            
            q_num = get_float(qty)
            p_num = get_float(price)
            t_num = get_float(total)
            
            # FILTRO ROBUSTO
            if p_num > 0 and desc_norm and "TOTAL" not in desc_norm:
                prod = {
                    'Descripcion': desc_val,
                    'Cantidad': q_num,
                    'Precio Unitario': p_num,
                    'Total Linea': t_num if t_num > 0 else (q_num * p_num)
                }
                products.append(prod)
                empty_streak = 0
                last_row_table = curr 
            else:
                if not desc_norm:
                    empty_streak += 1
                    if empty_streak > 15:
                        last_row_table = curr - 15
                        break
            curr += 1
    else:
        last_row_table = 20

    # --- OBSERVACIONES ---
    obs_text = None
    
    # 1. Intentar por etiqueta
    obs_label_pos = find_label_cell(sheet, ['OBSERVACIONES', 'REMARKS', 'NOTAS', 'GLOSA', 'COMENTARIOS'])
    if obs_label_pos:
        obs_text = extract_value_near_label(sheet, obs_label_pos[0], obs_label_pos[1], look_down=True, look_right=True)
    
    # 2. Si falló, fuerza bruta en el footer
    if not obs_text:
        obs_text = find_longest_text_in_footer(sheet, start_row=last_row_table + 1)
        
    # 3. Si falló, revisar Header/Footer de impresión
    if not obs_text:
        hf = scan_headers_footers(sheet)
        if hf and len(hf) > 20: 
             obs_text = hf

    header_data['Observaciones'] = obs_text

    # --- Unificación ---
    final_rows = []
    if products:
        for p in products:
            final_rows.append({'Archivo': uploaded_file.name, **header_data, **p})
    else:
        final_rows.append({'Archivo': uploaded_file.name, **header_data})
        
    return final_rows, None, debug_df, has_drawings

# --- Interfaz Gráfica ---

st.set_page_config(page_title="Extractor V15", layout="wide")
st.title("📄 Extractor de Facturas V15 (Prioridad Bajo Condición)")
st.markdown("""
**Ajustes V15:**
* **Prioridad Condición de Venta:** Busca primero frases como "BAJO CONDICION". Si las encuentra, ignora "COLLECT".
* **Depuración:** Detecta texto en encabezados y objetos flotantes.
""")

uploaded_files = st.file_uploader("Archivos Excel (.xlsx)", type=['xlsx'], accept_multiple_files=True)

if uploaded_files:
    if st.button("Procesar Archivos"):
        all_data = []
        debug_info = {}
        drawings_alert = []
        
        for file in uploaded_files:
            rows, err, debug_df, has_drawings = process_file(file)
            if has_drawings:
                drawings_alert.append(file.name)
            
            if rows: 
                all_data.extend(rows)
                debug_info[file.name] = debug_df
            if err: st.error(f"{file.name}: {err}")
        
        # Alerta de Cuadros de Texto
        if drawings_alert:
            st.warning(f"⚠️ ¡ATENCIÓN! Se detectaron objetos flotantes (Imágenes/Cuadros de Texto) en: {', '.join(drawings_alert)}. "
                       "Si faltan las Observaciones, es probable que estén dentro de estos cuadros. "
                       "**Solución:** Copia el texto del cuadro y pégalo en una celda vacía del Excel.")

        tab1, tab2 = st.tabs(["📊 Resultados", "🔍 Diagnóstico Total"])
        
        with tab1:
            if all_data:
                df = pd.DataFrame(all_data)
                cols_order = ['Archivo', 'Cliente', 'Num_Exp', 'Fecha', 'Condicion_Venta', 'Incoterm', 'Forma_Pago', 
                              'Moneda', 'Tipo_Cambio', 'Puerto_Emb', 'Puerto_Dest', 'Cantidad', 'Descripcion', 
                              'Precio Unitario', 'Total Linea', 'Observaciones']
                final_cols = [c for c in cols_order if c in df.columns] + [c for c in df.columns if c not in cols_order]
                df = df[final_cols]
                
                st.dataframe(df, use_container_width=True)
                
                buffer = io.BytesIO()
                with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False)
                st.download_button("Descargar Excel", buffer.getvalue(), "facturas_v15.xlsx")
            else:
                st.warning("No se extrajeron datos.")

        with tab2:
            st.info("Texto crudo visto por el sistema:")
            for fname, d_df in debug_info.items():
                with st.expander(f"Rayos X de: {fname}"):
                    st.dataframe(d_df, use_container_width=True)
