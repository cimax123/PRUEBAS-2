import streamlit as st
import pandas as pd
import openpyxl
from io import BytesIO
import datetime
import re

# Configuración de la página
st.set_page_config(page_title="Extractor de Facturas de Exportación", layout="wide")

def find_cell_by_keyword(sheet, keyword):
    """
    Busca una celda que contenga la palabra clave (case-insensitive).
    Retorna el objeto cell si lo encuentra, de lo contrario None.
    """
    keyword = keyword.lower()
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if keyword in cell.value.lower():
                    return cell
    return None

def get_value_near_keyword(sheet, keyword, search_radius=1):
    """
    Busca una palabra clave y devuelve el valor de la derecha o de abajo.
    Prioriza la celda de la derecha, luego la de abajo.
    """
    cell = find_cell_by_keyword(sheet, keyword)
    if not cell:
        return None
    
    # Intentar obtener valor a la derecha
    val_right = sheet.cell(row=cell.row, column=cell.column + 1).value
    if val_right is not None and str(val_right).strip() != "":
        return val_right
    
    # Intentar obtener valor abajo
    val_below = sheet.cell(row=cell.row + 1, column=cell.column).value
    return val_below

def extract_product_table(sheet):
    """
    Identifica la tabla de productos basada en encabezados y extrae filas
    hasta encontrar una fila vacía o una palabra de parada como 'TOTAL'.
    """
    header_cell = find_cell_by_keyword(sheet, "CANTIDAD") or find_cell_by_keyword(sheet, "QUANTITY")
    if not header_cell:
        return pd.DataFrame()

    # Identificar índices de columnas dinámicamente
    headers_row = header_cell.row
    col_map = {}
    
    # Escanear la fila de encabezados para mapear columnas
    for col in range(1, sheet.max_column + 1):
        val = sheet.cell(row=headers_row, column=col).value
        if not val: continue
        val_str = str(val).upper()
        
        if "CANT" in val_str or "QTY" in val_str: col_map['cantidad'] = col
        elif "DESC" in val_str: col_map['descripcion'] = col
        elif "UNIT" in val_str or "PRECIO" in val_str: col_map['precio_unitario'] = col
        elif "TOTAL" in val_str: col_map['total'] = col

    products = []
    curr_row = headers_row + 1
    
    # Extraer datos hasta encontrar el total o una fila vacía significativa
    while curr_row <= sheet.max_row:
        # Si la descripción o el total contienen la palabra "TOTAL", paramos
        desc_val = str(sheet.cell(row=curr_row, column=col_map.get('descripcion', 1)).value or "").upper()
        if "TOTAL" in desc_row_check(sheet, curr_row):
            break
            
        row_data = {
            "Cantidad": sheet.cell(row=curr_row, column=col_map.get('cantidad', 1)).value,
            "Descripción": sheet.cell(row=curr_row, column=col_map.get('descripcion', 1)).value,
            "Precio Unitario": sheet.cell(row=curr_row, column=col_map.get('precio_unitario', 1)).value,
            "Total": sheet.cell(row=curr_row, column=col_map.get('total', 1)).value,
        }
        
        # Si la fila está mayormente vacía, detenemos la extracción
        if not any(row_data.values()):
            break
            
        products.append(row_data)
        curr_row += 1
        
    return pd.DataFrame(products)

def desc_row_check(sheet, row_idx):
    """Verifica si alguna celda en la fila indica que es el final (Total)"""
    for col in range(1, 10):
        val = str(sheet.cell(row=row_idx, column=col).value or "").upper()
        if "TOTAL" in val: return "TOTAL"
    return ""

def process_invoice(file):
    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        sheet = wb.active
        
        # Extracción de Cabecera
        cliente = get_value_near_keyword(sheet, "CLIENTE") or get_value_near_keyword(sheet, "CUSTOMER")
        exp_num = get_value_near_keyword(sheet, "EXP")
        
        # Manejo de Fecha (Intentar extraer de una celda de fecha o texto)
        fecha_val = get_value_near_keyword(sheet, "FECHA") or get_value_near_keyword(sheet, "DATE")
        dia, mes, anio = None, None, None
        if isinstance(fecha_val, datetime.datetime):
            dia, mes, anio = fecha_val.day, fecha_val.month, fecha_val.year
        elif isinstance(fecha_val, str):
            # Intento básico de parseo si es string dd/mm/yyyy
            parts = re.findall(r'\d+', fecha_val)
            if len(parts) >= 3:
                dia, mes, anio = parts[0], parts[1], parts[2]

        # Logística
        incoterm = get_value_near_keyword(sheet, "INCOTERM") or get_value_near_keyword(sheet, "CONDICION")
        puerto_emb = get_value_near_keyword(sheet, "EMBARQUE") or get_value_near_keyword(sheet, "LOADING")
        puerto_dest = get_value_near_keyword(sheet, "DESTINO") or get_value_near_keyword(sheet, "DESTINATION")
        
        # Otros
        moneda = get_value_near_keyword(sheet, "MONEDA") or get_value_near_keyword(sheet, "CURRENCY")
        obs = get_value_near_keyword(sheet, "OBSERVACIONES") or get_value_near_keyword(sheet, "REMARKS")
        
        # Detalle de productos
        df_productos = extract_product_table(sheet)
        
        # Consolidar metadatos
        metadata = {
            "Cliente": cliente,
            "Número Exp": exp_num,
            "Año": anio,
            "Mes": mes,
            "Día": dia,
            "Incoterm": incoterm,
            "Puerto Embarque": puerto_emb,
            "Puerto Destino": puerto_dest,
            "Moneda": moneda,
            "Observaciones": obs
        }
        
        return metadata, df_productos

    except Exception as e:
        st.error(f"Error procesando el archivo: {e}")
        return None, None

# --- INTERFAZ DE USUARIO ---
st.title("📄 Procesador Inteligente de Facturas de Exportación")
st.markdown("""
Esta herramienta utiliza **lógica de proximidad** para extraer datos de Excel sin importar si las celdas cambian de posición ligeramente.
""")

uploaded_file = st.file_uploader("Carga tu archivo Excel (.xlsx)", type=["xlsx"])

if uploaded_file:
    with st.spinner("Analizando estructura del documento..."):
        metadata, df_items = process_invoice(uploaded_file)
        
    if metadata and not df_items.empty:
        st.success("¡Datos extraídos con éxito!")
        
        # Mostrar Cabecera
        st.subheader("Datos de Cabecera y Logística")
        col1, col2, col3 = st.columns(3)
        with col1:
            st.info(f"**Cliente:** {metadata['Cliente']}")
            st.info(f"**Nº Exp:** {metadata['Número Exp']}")
        with col2:
            st.info(f"**Fecha:** {metadata['Día']}/{metadata['Mes']}/{metadata['Año']}")
            st.info(f"**Incoterm:** {metadata['Incoterm']}")
        with col3:
            st.info(f"**Puerto Destino:** {metadata['Puerto Destino']}")
            st.info(f"**Moneda:** {metadata['Moneda']}")

        # Mostrar Tabla de Productos
        st.subheader("Detalle de Productos")
        st.dataframe(df_items, use_container_width=True)
        
        # Preparar descarga
        # Combinar metadata con productos para un reporte plano
        df_final = df_items.copy()
        for key, value in metadata.items():
            df_final[key] = value
            
        st.subheader("Exportar Resultados")
        col_ex1, col_ex2 = st.columns(2)
        
        # Botón CSV
        csv = df_final.to_csv(index=False).encode('utf-8-sig')
        col_ex1.download_button(
            label="Descargar como CSV",
            data=csv,
            file_name=f"Factura_{metadata['Número Exp'] or 'export'}.csv",
            mime='text/csv',
        )
        
        # Botón Excel
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_final.to_excel(writer, index=False, sheet_name='Datos_Extraidos')
        excel_data = output.getvalue()
        col_ex2.download_button(
            label="Descargar como Excel",
            data=excel_data,
            file_name=f"Factura_{metadata['Número Exp'] or 'export'}.xlsx",
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
    else:
        st.warning("No se pudo extraer la información necesaria. Verifica que el archivo contenga las palabras clave (CLIENTE, EXP, CANTIDAD, etc.)")

st.sidebar.header("Instrucciones")
st.sidebar.write("""
1. Sube un archivo .xlsx de factura.
2. El script buscará etiquetas como 'CLIENTE', 'EXP', 'FECHA' y 'CANTIDAD'.
3. Los datos deben estar a la derecha o justo debajo de la etiqueta.
4. La tabla de productos se extrae hasta encontrar la palabra 'TOTAL'.
""")
